import os
import numpy as np
import re
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd

# Path to the folder containing Excel files
input_path = "./toreplace/"

# Output folder for modified files
output_path = "./cleansed/"

# Read the replacement data from "replace.xlsx" using pandas
replace_array = np.array(pd.read_excel("replace.xlsx"))

# Custom function for case-insensitive find-and-replace
def custom_replace(cell_value, find_word, replace_word):
    return re.sub(r'\b' + find_word + r'\b', replace_word, cell_value, flags=re.IGNORECASE)

# Iterate through all Excel files in the input folder
for filename in os.listdir(input_path):
    if filename.endswith(".xlsx"):
        file_path = os.path.join(input_path, filename)
        
        # Print filename and timestamp when starting processing
        print(f"Processing file: {filename} ({datetime.now()})")
        
        # Load workbook and all sheet names
        workbook = load_workbook(file_path)
        sheet_names = workbook.sheetnames
        
        # Process each sheet
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            data = np.array([[cell.value if cell.value is not None else '' for cell in row] for row in sheet.iter_rows()])
            
            # Perform case-insensitive find-and-replace within sentences
            print("Applying replace for sheet: " + sheet_name)
            for row in replace_array:
                find_word = row[0]  # First column
                replace_word = row[1]  # Second column
                data = np.vectorize(custom_replace)(data.astype(str), find_word, replace_word)
            
            # Write modified data back to the sheet
            print("Writing data back to the sheet")
            for i, row in enumerate(sheet.iter_rows()):
                for j, cell in enumerate(row):
                    cell.value = data[i, j]
        
        # Save the modified workbook to a new file in the output folder
        print("Creating new workbook.")
        output_file_path = os.path.join(output_path, f"modified_{filename}")
        workbook.save(output_file_path)

        # Print completion message
        print(f"Processing completed for file: {filename} ({datetime.now()})")

print("Find-and-replace completed for all Excel files. Modified files saved with 'modified_' prefix in " + output_path)
